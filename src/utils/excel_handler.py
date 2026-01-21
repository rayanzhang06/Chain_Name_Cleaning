"""
Excel 处理工具

提供 Excel 文件的读取、写入、格式保留等功能。
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Union

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Excel 处理器 - 封装 Excel 读写操作

    功能：
    - 读取 Excel 文件
    - 写入 Excel 文件（保留格式）
    - 数据预处理
    - 批量处理
    """

    def __init__(self, file_path: Optional[Path] = None):
        """
        初始化 Excel 处理器

        Args:
            file_path: Excel 文件路径（可选）
        """
        self.file_path = Path(file_path) if file_path else None
        self.workbook = None
        self.worksheet = None

    def read_excel(
        self,
        file_path: Union[str, Path],
        sheet_name: Union[str, int] = 0,
        header: int = 0,
        **kwargs
    ) -> pd.DataFrame:
        """
        读取 Excel 文件

        Args:
            file_path: 文件路径
            sheet_name: 工作表名称或索引
            header: 表头行索引
            **kwargs: 传递给 pd.read_excel 的其他参数

        Returns:
            DataFrame
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header,
                **kwargs
            )

            # 去除列名前后空格
            df.columns = df.columns.str.strip()

            logger.info(f"读取 Excel 文件: {file_path}, 形状: {df.shape}")
            return df

        except Exception as e:
            logger.error(f"读取 Excel 失败: {file_path}, 错误: {e}")
            raise

    def write_excel(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        sheet_name: str = 'Sheet1',
        index: bool = False,
        preserve_format: bool = False,
        **kwargs
    ) -> None:
        """
        写入 Excel 文件

        Args:
            df: DataFrame
            output_path: 输出路径
            sheet_name: 工作表名称
            index: 是否写入索引
            preserve_format: 是否保留格式（从原文件）
            **kwargs: 传递给 df.to_excel 的其他参数
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            if preserve_format and self.file_path and self.file_path.exists():
                # 保留格式：使用 openpyxl
                self._write_with_format(df, output_path, sheet_name, index)
            else:
                # 直接写入
                df.to_excel(output_path, sheet_name=sheet_name, index=index, **kwargs)

            logger.info(f"写入 Excel 文件: {output_path}, 行数: {len(df)}")

        except Exception as e:
            logger.error(f"写入 Excel 失败: {output_path}, 错误: {e}")
            raise

    def _write_with_format(
        self,
        df: pd.DataFrame,
        output_path: Path,
        sheet_name: str,
        index: bool
    ) -> None:
        """
        写入 Excel 并保留格式

        Args:
            df: DataFrame
            output_path: 输出路径
            sheet_name: 工作表名称
            index: 是否写入索引
        """
        # 如果原文件存在，复制原文件
        if self.file_path and self.file_path.exists():
            import shutil
            shutil.copy(self.file_path, output_path)

        # 加载工作簿
        wb = openpyxl.load_workbook(output_path)

        # 删除或创建工作表
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name, 0)

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # 应用格式
                if r_idx == 1:  # 表头
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存
        wb.save(output_path)

    def preprocess_dataframe(
        self,
        df: pd.DataFrame,
        drop_duplicates: bool = True,
        drop_empty: bool = True,
        normalize_text: bool = True,
        strip_whitespace: bool = True
    ) -> pd.DataFrame:
        """
        预处理 DataFrame

        Args:
            df: DataFrame
            drop_duplicates: 是否删除重复行
            drop_empty: 是否删除空行
            normalize_text: 是否标准化文本（去除特殊字符）
            strip_whitespace: 是否去除空格

        Returns:
            处理后的 DataFrame
        """
        df_processed = df.copy()

        # 去除列名空格
        if strip_whitespace:
            df_processed.columns = df_processed.columns.str.strip()

        # 去除数据空格
        if strip_whitespace:
            df_processed = df_processed.apply(
                lambda col: col.str.strip() if col.dtype == 'object' else col
            )

        # 删除重复行
        if drop_duplicates:
            before_count = len(df_processed)
            df_processed = df_processed.drop_duplicates()
            after_count = len(df_processed)
            if before_count > after_count:
                logger.info(f"删除重复行: {before_count - after_count} 行")

        # 删除空行
        if drop_empty:
            before_count = len(df_processed)
            df_processed = df_processed.dropna(how='all')
            after_count = len(df_processed)
            if before_count > after_count:
                logger.info(f"删除空行: {before_count - after_count} 行")

        # 标准化文本
        if normalize_text:
            # 移除不可见字符
            df_processed = df_processed.replace(
                to_replace=[r'\\r\\n', r'\\r', r'\\n', r'\\t'],
                value='',
                regex=True
            )

        logger.info(f"预处理完成: {df_processed.shape}")
        return df_processed

    def get_sheet_names(self, file_path: Optional[Union[str, Path]] = None) -> List[str]:
        """
        获取工作表名称列表

        Args:
            file_path: 文件路径（可选，默认使用初始化时的路径）

        Returns:
            工作表名称列表
        """
        file_path = Path(file_path) if file_path else self.file_path

        if not file_path or not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names

        except Exception as e:
            logger.error(f"获取工作表名称失败: {file_path}, 错误: {e}")
            raise

    def batch_read_excel(
        self,
        file_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None,
        **kwargs
    ) -> Dict[str, pd.DataFrame]:
        """
        批量读取多个工作表

        Args:
            file_path: 文件路径
            sheet_names: 工作表名称列表（可选，默认读取所有）
            **kwargs: 传递给 read_excel 的参数

        Returns:
            {工作表名: DataFrame} 字典
        """
        file_path = Path(file_path)

        if sheet_names is None:
            sheet_names = self.get_sheet_names(file_path)

        result = {}
        for sheet_name in sheet_names:
            try:
                df = self.read_excel(file_path, sheet_name=sheet_name, **kwargs)
                result[sheet_name] = df
            except Exception as e:
                logger.error(f"读取工作表失败: {sheet_name}, 错误: {e}")

        logger.info(f"批量读取完成: {len(result)} 个工作表")
        return result

    def export_multiple_sheets(
        self,
        data_dict: Dict[str, pd.DataFrame],
        output_path: Union[str, Path],
        index: bool = False
    ) -> None:
        """
        导出多个工作表到单个 Excel 文件

        Args:
            data_dict: {工作表名: DataFrame} 字典
            output_path: 输出路径
            index: 是否写入索引
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=index)

            logger.info(f"导出多工作表 Excel: {output_path}, 工作表数: {len(data_dict)}")

        except Exception as e:
            logger.error(f"导出失败: {output_path}, 错误: {e}")
            raise


def quick_read_excel(file_path: Union[str, Path], **kwargs) -> pd.DataFrame:
    """
    快速读取 Excel 文件（便捷函数）

    Args:
        file_path: 文件路径
        **kwargs: 传递给 read_excel 的参数

    Returns:
        DataFrame
    """
    handler = ExcelHandler()
    return handler.read_excel(file_path, **kwargs)


def quick_write_excel(
    df: pd.DataFrame,
    output_path: Union[str, Path],
    **kwargs
) -> None:
    """
    快速写入 Excel 文件（便捷函数）

    Args:
        df: DataFrame
        output_path: 输出路径
        **kwargs: 传递给 write_excel 的参数
    """
    handler = ExcelHandler()
    handler.write_excel(df, output_path, **kwargs)
